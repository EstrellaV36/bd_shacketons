import pandas as pd
from PyQt6.QtWidgets import QWidget, QVBoxLayout, QLabel, QPushButton, QHBoxLayout, QTableWidget, QTableWidgetItem, QComboBox, QFileDialog, QDateEdit, QCheckBox
from PyQt6.QtCore import Qt, QDate
from app.database import get_db_session
from app.models import Buque, EtaCiudad, Tripulante, Vuelo, TripulanteVuelo, Restaurante, TripulanteRestaurante, Transporte, TripulanteTransporte, Hotel, TripulanteHotel, Buque, TripulanteAsistencia
from app.controllers import CITY_AIRPORT_CODES, CITY_TO_AIRPORT_CODES
from openpyxl.styles import PatternFill
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
from datetime import datetime, time, timedelta
from sqlalchemy import func, and_, or_
from collections import defaultdict


class GeneracionReportesScreen(QWidget):
    def __init__(self, main_window):
        super().__init__()
        self.main_window = main_window
        self.setup_ui()

    def setup_ui(self):
        layout = QVBoxLayout(self)

        # Añadir un botón "Volver" al menú principal
        self.button_volver = QPushButton("Volver")
        self.button_volver.setFixedWidth(100)
        self.button_volver.clicked.connect(self.volver_al_menu_principal)
        layout.addWidget(self.button_volver, alignment=Qt.AlignmentFlag.AlignLeft)

        # Crear un layout horizontal para los botones
        layout_botones = QHBoxLayout()
        layout_botones.setAlignment(Qt.AlignmentFlag.AlignCenter)

        # Botones
        button_informar = QPushButton("Informar")
        button_informar.setFixedSize(140, 40)
        layout_botones.addWidget(button_informar)

        button_programar = QPushButton("Programar")
        button_programar.setFixedSize(140, 40)
        button_programar.clicked.connect(self.mostrar_opciones_programar)
        layout_botones.addWidget(button_programar)

        button_liquidar = QPushButton("Liquidar")
        button_liquidar.setFixedSize(140, 40)
        layout_botones.addWidget(button_liquidar)

        button_cuadrar_proveedor = QPushButton("Cuadrar Proveedor")
        button_cuadrar_proveedor.setFixedSize(140, 40)
        layout_botones.addWidget(button_cuadrar_proveedor)

        # Añadir el layout de botones al layout principal
        layout.addLayout(layout_botones)

    def mostrar_opciones_programar(self):
        opciones_programar_screen = OpcionesProgramarScreen(self.main_window)
        # Guardar el índice de OpcionesProgramarScreen
        self.main_window.opciones_programar_index = self.main_window.stacked_widget.addWidget(opciones_programar_screen)
        self.main_window.stacked_widget.setCurrentWidget(opciones_programar_screen)

    def volver_al_menu_principal(self):
        self.main_window.stacked_widget.setCurrentIndex(0)  # Regresar al menú principal

class OpcionesProgramarScreen(QWidget):
    def __init__(self, main_window):
        super().__init__()
        self.main_window = main_window
        self.setup_ui()
        self.menu_reportes_index = main_window.stacked_widget.addWidget(self)

    def setup_ui(self):
        layout = QVBoxLayout(self)

        # Botón "Volver" para regresar a la pantalla anterior (Generación de Reportes)
        button_volver = QPushButton("Volver")
        button_volver.setFixedWidth(100)
        button_volver.clicked.connect(self.volver_a_reportes)
        layout.addWidget(button_volver, alignment=Qt.AlignmentFlag.AlignLeft)

        layout_botones = QHBoxLayout()
        layout_botones.setAlignment(Qt.AlignmentFlag.AlignCenter)

        # Botones
        button_asistencias = QPushButton("Asistencias")
        button_asistencias.setFixedSize(140, 40)
        button_asistencias.clicked.connect(self.mostrar_asistencias)
        layout_botones.addWidget(button_asistencias)

        button_transportes = QPushButton("Req. transportes")
        button_transportes.setFixedSize(140, 40)
        button_transportes.clicked.connect(self.mostrar_transportes)
        layout_botones.addWidget(button_transportes)

        button_room_list = QPushButton("Room List")
        button_room_list.setFixedSize(140, 40)
        button_room_list.clicked.connect(self.mostrar_room_list)
        layout_botones.addWidget(button_room_list)

        button_hoteles = QPushButton("Req. Hoteles")
        button_hoteles.setFixedSize(140, 40)
        button_hoteles.clicked.connect(self.mostrar_req_hoteles)
        layout_botones.addWidget(button_hoteles)

        # Añadir el layout de botones al layout principal
        layout.addLayout(layout_botones)

    def mostrar_asistencias(self):
        asistencias_screen = AsistenciasScreen(self.main_window)
        asistencias_screen.opciones_programar_index = self.main_window.opciones_programar_index
        self.main_window.stacked_widget.addWidget(asistencias_screen)
        self.main_window.stacked_widget.setCurrentWidget(asistencias_screen)

    def mostrar_transportes(self):
        transportes_screen = TransportesScreen(self.main_window)
        self.main_window.stacked_widget.addWidget(transportes_screen)
        self.main_window.stacked_widget.setCurrentWidget(transportes_screen)

    def mostrar_room_list(self):
        room_list_screen = RoomListScreen(self.main_window)
        self.main_window.stacked_widget.addWidget(room_list_screen)
        self.main_window.stacked_widget.setCurrentWidget(room_list_screen)

    def mostrar_req_hoteles(self):
        hotel_screen = HotelScreen(self.main_window)
        self.main_window.stacked_widget.addWidget(hotel_screen)
        self.main_window.stacked_widget.setCurrentWidget(hotel_screen)

    def volver_a_reportes(self):
        self.main_window.stacked_widget.setCurrentIndex(self.main_window.menu_reportes_index)

class HotelScreen(QWidget):
    def __init__(self, main_window):
        super().__init__()
        self.main_window = main_window
        self.setup_ui()

    def setup_ui(self):
        layout = QVBoxLayout(self)

        self.combo_ciudades = QComboBox()
        self.combo_ciudades.addItems(["SCL", "PUQ", "WPU"]) 
        layout.addWidget(self.combo_ciudades)

        # Cuadro de selección de buque
        self.combo_buques = QComboBox()
        self.combo_buques.addItems(["Buque", "Silver Endeavour", "C-GEAI", "C-FMKB", "Silver Cloud", "Fram", "SYLVIA EARLE"])
        layout.addWidget(self.combo_buques)

        # Filtro por fechas
        self.check_fecha = QCheckBox("Habilitar filtro por fechas")
        self.check_fecha.setChecked(False)
        self.check_fecha.stateChanged.connect(self.actualizar_datos)
        layout.addWidget(self.check_fecha)

        # Layout horizontal para las fechas
        layout_filtro = QHBoxLayout()
        self.date_start1 = QDateEdit()
        self.date_start1.setCalendarPopup(True)
        self.date_start1.setDate(QDate.currentDate())
        layout_filtro.addWidget(QLabel("Fecha inicio:"))
        layout_filtro.addWidget(self.date_start1)

        self.date_end1 = QDateEdit()
        self.date_end1.setCalendarPopup(True)
        self.date_end1.setDate(QDate.currentDate())
        layout_filtro.addWidget(QLabel("Fecha fin:"))
        layout_filtro.addWidget(self.date_end1)

        layout.addLayout(layout_filtro)

        self.label = QLabel()
        layout.addWidget(self.label)

        # Tabla para mostrar los datos
        self.table_widget = QTableWidget()
        layout.addWidget(self.table_widget)

        # Botón para generar el Excel
        button_generar_excel = QPushButton("Generar Excel")
        button_generar_excel.clicked.connect(self.generar_excel_con_ciudad)
        layout.addWidget(button_generar_excel)

        # Botón para volver a la pantalla principal
        button_volver = QPushButton("Volver")
        button_volver.setFixedWidth(100)
        button_volver.clicked.connect(self.volver_a_opciones_programar)
        layout.addWidget(button_volver, alignment=Qt.AlignmentFlag.AlignLeft)

        # Conectar señales
        self.combo_ciudades.currentTextChanged.connect(self.actualizar_datos)
        self.combo_buques.currentTextChanged.connect(self.actualizar_datos)
        self.date_start1.dateChanged.connect(self.actualizar_datos)
        self.date_end1.dateChanged.connect(self.actualizar_datos)

        # Actualizar los datos inicialmente
        self.actualizar_datos()

    def generar_excel_con_ciudad(self):
        ciudad_seleccionada = self.combo_ciudades.currentText()
        buque_seleccionado = self.combo_buques.currentText()
        # Obtener las fechas seleccionadas
        fecha_inicio = self.date_start1.date().toString("dd-MM-yyyy")  # Formato de fecha: dd-mm-aaaa
        fecha_fin = self.date_end1.date().toString("dd-MM-yyyy")

    # Llamar a la función generar_excel con ciudad, buque y las fechas
        self.generar_excel(ciudad_seleccionada, buque_seleccionado, fecha_inicio, fecha_fin)

    def actualizar_datos(self):
        ciudad_seleccionada = self.combo_ciudades.currentText()
        buque_seleccionado = self.combo_buques.currentText()
        self.label.setText(f"Lista de hoteles en {ciudad_seleccionada}")

        # Cargar datos basados en los filtros seleccionados
        self.cargar_datos(ciudad_seleccionada, buque_seleccionado)

    def cargar_datos(self, ciudad_seleccionada, buque_seleccionado):
        session = get_db_session()
        ciudad_seleccionada = CITY_AIRPORT_CODES.get(ciudad_seleccionada)

        buque_seleccionado = buque_seleccionado.lower()

        ciudad_select = CITY_TO_AIRPORT_CODES.get(ciudad_seleccionada)

        # Obtener fechas seleccionadas
        fecha_inicio = self.date_start1.date().toPyDate()
        fecha_fin = datetime.combine(self.date_end1.date().toPyDate(), time.max)

        hotel_necesario = (
            session.query(
                Tripulante.tripulante_id,
                Tripulante.estado.label("Estado"),
                Tripulante.nombre.label("First_Name"),
                Tripulante.apellido.label("Last_Name"),
                Tripulante.sexo.label("Genero"),
                Tripulante.nacionalidad.label("Nacionalidad"),
                Tripulante.posicion.label("Position"),
                TripulanteHotel.categoria.label("Categoria"),
                Hotel.ciudad.label("Ciudad_Hotel"),
                Hotel.nombre.label("Nombre_Hotel"),
                TripulanteHotel.fecha_entrada.label("check_in"),
                TripulanteHotel.fecha_salida.label("check_out"),
                TripulanteHotel.tipo_habitacion.label("Rooms")
            )
            .join(TripulanteHotel, Tripulante.tripulante_id == TripulanteHotel.tripulante_id)
            .join(Buque, Tripulante.buque_id == Buque.buque_id)
            .join(Hotel, Hotel.hotel_id == TripulanteHotel.hotel_id)
            .filter(or_(
                func.lower(Hotel.ciudad) == f"hotel {ciudad_select.lower()}",
                func.lower(Hotel.ciudad) == f"day use {ciudad_select.lower()}"
            ))
        )

        for tripulante in hotel_necesario:
            print(f"Nombre: {tripulante.First_Name}, Check-in: {tripulante.check_in}, Check-out: {tripulante.check_out}")

        if buque_seleccionado != "buque":
            hotel_necesario = hotel_necesario.filter(func.lower(Buque.nombre) == buque_seleccionado)

        if self.check_fecha.isChecked():
            # Obtener las fechas seleccionadas de la interfaz
            fecha_inicio_date = datetime.combine(fecha_inicio, time.min)  # Primer momento del día
            fecha_fin_date = datetime.combine(fecha_fin, time.max)  # Último momento del día

            # Obtener las fechas ETA y ETD del tripulante asociado al buque seleccionado
            fechas_eta_etd = (
                session.query(EtaCiudad.eta, EtaCiudad.etd)
                .join(Tripulante, EtaCiudad.tripulante_id == Tripulante.tripulante_id)  
                .join(Buque, Tripulante.buque_id == Buque.buque_id) 
                .filter(func.lower(Buque.nombre) == buque_seleccionado)
                .filter(EtaCiudad.eta >= fecha_inicio_date, EtaCiudad.eta <= fecha_fin_date)
                .first()
            )

            if fechas_eta_etd:
                eta = fechas_eta_etd.eta.date()  # Convertir a date
                etd = fechas_eta_etd.etd.date()  # Convertir a date

                # Filtrar por las fechas de entrada y salida del hotel
                hotel_necesario = hotel_necesario.filter(
                    TripulanteHotel.fecha_entrada >= eta,
                    TripulanteHotel.fecha_salida <= etd
                )

        # Obtener los resultados de la consulta
        resultados = hotel_necesario.all()

        # Limpiar tabla y mostrar resultados
        self.table_widget.setRowCount(0)
        self.table_widget.setColumnCount(10)
        self.table_widget.setHorizontalHeaderLabels(
            ["First Name", "Last Name", "Gender", "Nacionalidad", "Position", "Categoria", "Hotel Ciudad", "Check In", "Check Out", "Rooms"]
        )

        # Rellenar la tabla con los resultados
        for resultado in resultados:
            row = self.table_widget.rowCount()
            self.table_widget.insertRow(row)

            # Asumiendo que resultado es una tupla y que los índices coinciden con los campos de la consulta
            self.table_widget.setItem(row, 0, QTableWidgetItem(resultado.First_Name))  # First Name
            self.table_widget.setItem(row, 1, QTableWidgetItem(resultado.Last_Name))   # Last Name
            self.table_widget.setItem(row, 2, QTableWidgetItem(resultado.Genero))      # Gender
            self.table_widget.setItem(row, 3, QTableWidgetItem(resultado.Nacionalidad)) # Nacionalidad
            self.table_widget.setItem(row, 4, QTableWidgetItem(resultado.Position))     # Position
            self.table_widget.setItem(row, 5, QTableWidgetItem(str(resultado.Categoria))) # Categoria
            self.table_widget.setItem(row, 6, QTableWidgetItem(resultado.Ciudad_Hotel)) # Hotel Ciudad
            self.table_widget.setItem(row, 7, QTableWidgetItem(str(resultado.check_in)))   # Check In
            self.table_widget.setItem(row, 8, QTableWidgetItem(str(resultado.check_out)))  # Check Out
            self.table_widget.setItem(row, 9, QTableWidgetItem(resultado.Rooms))  # Rooms

    def generar_excel(self, hotel_seleccionado, buque_seleccionado, fecha_inicio, fecha_fin):
        # Crear un DataFrame con los datos de la tabla
        data = []
        for row in range(self.table_widget.rowCount()):
            row_data = []
            for column in range(self.table_widget.columnCount()):
                item = self.table_widget.item(row, column)
                row_data.append(item.text() if item else "")
            data.append(row_data)

        # Definir los nombres de las columnas
        column_names = ["First name", "Last name", "Gender", "Nacionalidad", "Position", 
                        "Categoria", "Hotel Ciudad", "Check in", "Check out", "Rooms"]

        df = pd.DataFrame(data, columns=column_names)

        # Ordenar el DataFrame por "Categoria" y "Check in"
        df.sort_values(by=["Categoria", "Check in"], ascending=[True, True], inplace=True)

        df.insert(0, "Nro", range(1, len(df) + 1))

        # Generar nombre de archivo basado en buque y hotel seleccionados
        file_name_parts = ["informe_hotel"]
        if hotel_seleccionado.lower() != "hotel" and hotel_seleccionado.lower() not in file_name_parts:
            file_name_parts.append(hotel_seleccionado)
        if buque_seleccionado.lower() != "buque" and buque_seleccionado.lower() not in file_name_parts:
            file_name_parts.append(buque_seleccionado)

        file_name = "_".join(file_name_parts) + ".xlsx"

        # Guardar archivo Excel
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Guardar archivo Excel",
            file_name,
            "Excel Files (*.xlsx);;All Files (*)"
        )

        if file_path:
            wb = Workbook()
            ws = wb.active

            # Encabezados personalizados
            ws.cell(row=1, column=1, value="BUQUE").font = Font(bold=True)
            ws.cell(row=1, column=2, value=buque_seleccionado if buque_seleccionado else "N/A")

            ws.cell(row=2, column=1, value="CIUDAD").font = Font(bold=True)
            ws.cell(row=2, column=2, value=hotel_seleccionado if hotel_seleccionado else "N/A")

            # Solo mostrar las fechas si están seleccionadas
            if fecha_inicio and fecha_fin:
                ws.cell(row=3, column=1, value="ETA Desde:").font = Font(bold=True)
                ws.cell(row=3, column=2, value=fecha_inicio)
                ws.cell(row=4, column=1, value="ETA Hasta:").font = Font(bold=True)
                ws.cell(row=4, column=2, value=fecha_fin)
            else:
                ws.cell(row=3, column=1, value="ETA No seleccionada")

            # Escribir el texto final antes de la tabla
            ws.cell(row=6, column=1, value="Informe necesidad hotel").font = Font(bold=True)

            # Aplicar negrita a los encabezados
            header_fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')
            for col_num, col_name in enumerate(df.columns, 1):
                cell = ws.cell(row=7, column=col_num)  # Cambiar a fila 7
                cell.value = col_name
                cell.fill = header_fill
                cell.font = Font(bold=True)  # Negrita

            # Escribir los datos del DataFrame
            data_fill = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')
            for row_num, row_data in enumerate(df.values, start=8):  # Cambiar a fila 8
                for col_num, cell_value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.value = cell_value
                    cell.fill = data_fill

            # Centrar el número correlativo
            for row_num in range(8, len(df) + 8):  # Cambiar el rango según donde se escriban los datos
                ws.cell(row=row_num, column=1).alignment = Alignment(horizontal='center')

            # Ajustar el ancho de las columnas
            for col_num in range(1, len(df.columns) + 1):
                max_length = 0
                column = get_column_letter(col_num)
                for row in ws[column]:
                    try:
                        if len(str(row.value)) > max_length:
                            max_length = len(str(row.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)  # Agregar un poco de espacio extra
                ws.column_dimensions[column].width = adjusted_width

            # Guardar el archivo Excel
            wb.save(file_path)

    def volver_a_opciones_programar(self):
        self.main_window.stacked_widget.setCurrentIndex(self.main_window.menu_reportes_index)

class RoomListScreen(QWidget):
    def __init__(self, main_window):
        super().__init__()
        self.main_window = main_window
        self.setup_ui()

    def setup_ui(self):
        layout = QVBoxLayout(self)

        # Cuadro de selección de hotel
        self.combo_hoteles = QComboBox()
        self.combo_hoteles.addItems(["Hotel", "CABO DE HORNOS", "TBC", "DIEGO DE ALMAGRO", "SIN HOTEL", "Holiday Inn", "Lukataia"])
        layout.addWidget(self.combo_hoteles)

        # Cuadro de selección de buque
        self.combo_buques = QComboBox()
        self.combo_buques.addItems(["Buque", "Silver Endeavour", "C-GEAI", "C-FMKB", "Silver Cloud", "Fram", "SYLVIA EARLE"])
        layout.addWidget(self.combo_buques)

        self.check_fecha = QCheckBox("Habilitar filtro por fechas")
        self.check_fecha.setChecked(False)  # Inicialmente deshabilitado
        self.check_fecha.stateChanged.connect(self.actualizar_datos)  # Conectar evento de cambio de estado
        layout.addWidget(self.check_fecha)

        # Layout horizontal para las fechas
        layout_filtro = QHBoxLayout()

        # Selector de fecha de inicio
        self.date_start1 = QDateEdit()
        self.date_start1.setCalendarPopup(True)
        self.date_start1.setDate(QDate.currentDate())
        layout_filtro.addWidget(QLabel("Fecha inicio:"))
        layout_filtro.addWidget(self.date_start1)

        # Selector de fecha de fin
        self.date_end1 = QDateEdit()
        self.date_end1.setCalendarPopup(True)
        self.date_end1.setDate(QDate.currentDate())
        layout_filtro.addWidget(QLabel("Fecha fin:"))
        layout_filtro.addWidget(self.date_end1)

        # Agregar el layout horizontal al layout principal
        layout.addLayout(layout_filtro)

        # Label para mostrar asistencias
        self.label = QLabel()
        layout.addWidget(self.label)

        # Tabla para mostrar datos
        self.table_widget = QTableWidget()
        layout.addWidget(self.table_widget)

        # Botón para generar el Excel
        button_generar_excel = QPushButton("Generar Excel")
        button_generar_excel.clicked.connect(self.generar_excel_con_ciudad)
        layout.addWidget(button_generar_excel)

        # Botón "Volver" para regresar a la pantalla anterior
        button_volver = QPushButton("Volver")
        button_volver.setFixedWidth(100)
        button_volver.clicked.connect(self.volver_a_opciones_programar)
        layout.addWidget(button_volver, alignment=Qt.AlignmentFlag.AlignLeft)

        # Conectar cambios en los QComboBox
        self.combo_hoteles.currentTextChanged.connect(self.actualizar_datos)
        self.combo_buques.currentTextChanged.connect(self.actualizar_datos)
        self.date_start1.dateChanged.connect(self.actualizar_datos)
        self.date_end1.dateChanged.connect(self.actualizar_datos)

        # Actualizar datos inicialmente
        self.actualizar_datos()

    def generar_excel_con_ciudad(self):
        hotel_seleccionado = self.combo_hoteles.currentText()  # Obtener la ciudad seleccionada
        buque_seleccionado = self.combo_buques.currentText()  # Obtener la ciudad seleccionada
        self.generar_excel(hotel_seleccionado, buque_seleccionado)  # Llamar a generar_excel con la ciudad seleccionada

    def actualizar_datos(self):
        hotel_seleccionado = self.combo_hoteles.currentText()  # Obtiene la ciudad seleccionada
        buque_seleccionado = self.combo_buques.currentText()  # Obtiene la ciudad seleccionada
        self.label.setText(f"Room list en {hotel_seleccionado}")  # Actualiza el label

        # Cargar datos en la tabla
        self.cargar_datos(hotel_seleccionado, buque_seleccionado)

    def cargar_datos(self, hotel_seleccionado, buque_seleccionado):
        session = get_db_session()  # Obtener la sesión de la base de datos

        # Convertir las entradas a minúsculas para comparación
        hotel_seleccionado = hotel_seleccionado.lower()
        buque_seleccionado = buque_seleccionado.lower()

        # Obtener las fechas seleccionadas en QDateEdit
        fecha_inicio = self.date_start1.date().toPyDate()  # Convertir a objeto de fecha de Python
        fecha_fin = datetime.combine(self.date_end1.date().toPyDate(), time.max)  # Combinar con la hora máxima del día

        # Construir la consulta de roomlist
        roomlist_query = (
            session.query(
                Hotel.nombre.label("Nombre_hotel"),
                TripulanteHotel.fecha_entrada.label("check_in"),
                TripulanteHotel.fecha_salida.label("check_out"),
                TripulanteHotel.tipo_habitacion.label("Rooms"),
                Tripulante.nombre.label("First_Name"),
                Tripulante.apellido.label("Last_Name"),
                Tripulante.sexo.label("Gender"),
                Tripulante.nacionalidad.label("Nacionalidad"),
                Tripulante.posicion.label("Position"),
                Tripulante.tripulante_id
            )
            .join(Buque, Buque.buque_id == Tripulante.buque_id)
            .filter(Tripulante.buque_id == EtaCiudad.buque_id)
            .filter(Tripulante.tripulante_id == EtaCiudad.tripulante_id)
            .filter(Tripulante.tripulante_id == TripulanteHotel.tripulante_id)
            .filter(TripulanteHotel.hotel_id == Hotel.hotel_id)
            .distinct()
        )

        # Aplicar filtro de hotel si se seleccionó uno específico
        if hotel_seleccionado != "hotel":
            roomlist_query = roomlist_query.filter(func.lower(Hotel.nombre) == hotel_seleccionado)

        # Aplicar filtro de buque si se seleccionó uno específico
        if buque_seleccionado != "buque":
            roomlist_query = roomlist_query.filter(func.lower(Buque.nombre) == buque_seleccionado)

        # Aplicar filtro de ETA por rango de fechas si está habilitado
        if self.check_fecha.isChecked():
            roomlist_query = (
                roomlist_query
                .join(EtaCiudad, and_(
                    Tripulante.buque_id == EtaCiudad.buque_id,
                    EtaCiudad.eta >= fecha_inicio,
                    EtaCiudad.eta <= fecha_fin
                ))
                .filter(
                    EtaCiudad.eta >= fecha_inicio,
                    EtaCiudad.eta <= fecha_fin
                )
            )

        # Limpiar la tabla
        self.table_widget.setRowCount(0)
        self.table_widget.setColumnCount(8)  # Número correcto de columnas
        self.table_widget.setHorizontalHeaderLabels(
            ["First Name", "Last Name", "Gender", "Nacionalidad", "Position", "Check In", "Check Out", "Rooms"]
        )

        # Llenar la tabla con los resultados de la consulta
        for roomlist in roomlist_query:
            row = self.table_widget.rowCount()
            self.table_widget.insertRow(row)
            
            self.table_widget.setItem(row, 0, QTableWidgetItem(str(roomlist.First_Name)))  # First Name
            self.table_widget.setItem(row, 1, QTableWidgetItem(str(roomlist.Last_Name)))  # Last Name
            self.table_widget.setItem(row, 2, QTableWidgetItem(str(roomlist.Gender)))  # Gender
            self.table_widget.setItem(row, 3, QTableWidgetItem(str(roomlist.Nacionalidad)))  # Nacionalidad
            self.table_widget.setItem(row, 4, QTableWidgetItem(str(roomlist.Position)))  # Position
            
            # Check In
            self.table_widget.setItem(row, 5, QTableWidgetItem(str(roomlist.check_in) if roomlist.check_in else ""))
            # Check Out
            self.table_widget.setItem(row, 6, QTableWidgetItem(str(roomlist.check_out) if roomlist.check_out else ""))
            # Rooms
            self.table_widget.setItem(row, 7, QTableWidgetItem(str(roomlist.Rooms) if roomlist.Rooms else ""))

    def generar_excel(self, hotel_seleccionado, buque_seleccionado):
        def incrementar_grupo(group_counter):
            group_list = list(group_counter)

            i = len(group_list) - 1

            while i >= 0:
                if group_list[i] != 'Z':
                    group_list[i] = chr(ord(group_list[i]) + 1)
                    return ''.join(group_list)
                else:
                    group_list[i] = 'A'
                    i -= 1

            return 'A' + ''.join(group_list)
            
        data = []

        for row in range(self.table_widget.rowCount()):
            row_data = []
            for column in range(self.table_widget.columnCount()):
                item = self.table_widget.item(row, column)
                row_data.append(item.text() if item else "")
            data.append(row_data)

        column_names = []
        for i in range(self.table_widget.columnCount()):
            header_item = self.table_widget.horizontalHeaderItem(i)
            if header_item is not None:
                column_names.append(header_item.text())
            else:
                column_names.append(f"Column {i + 1}")

        df = pd.DataFrame(data, columns=column_names)

        # Convertir 'Check In' a datetime
        df['Check In'] = pd.to_datetime(df['Check In'], errors='coerce')

        # Ordenar el DataFrame primero por 'Check In', luego por 'Position' y finalmente por 'Gender'
        df = df.sort_values(by=["Check In", "Position", "Gender"])

        # Formatear 'Check In' de nuevo a string con el formato deseado
        df['Check In'] = df['Check In'].dt.strftime('%Y-%m-%d')

        
        df.insert(0, "Nro", range(1, len(df) + 1))

        group_counter = 'A'
        double_buffer = []

        for idx, row in df.iterrows():
            room_type = row['Rooms']
            
            if "Doble" in room_type:
                double_buffer.append(idx)

                if len(double_buffer) == 1:
                    group_aux = group_counter
                    df.loc[double_buffer, 'Grupo'] = group_aux
                    group_counter = incrementar_grupo(group_counter)
                
                if len(double_buffer) == 2:
                    df.loc[double_buffer, 'Grupo'] = group_aux
                    double_buffer = []

            elif "Single" in room_type:
                df.loc[idx, 'Grupo'] = group_counter
                group_counter = incrementar_grupo(group_counter)

            else:
                df.loc[idx, 'Grupo'] = ""

        if len(double_buffer) == 1:
            df.loc[double_buffer, 'Grupo'] = group_counter

        file_name_parts = ["room_list"]
        if hotel_seleccionado.lower() != "hotel" and hotel_seleccionado.lower() not in file_name_parts:
            file_name_parts.append(hotel_seleccionado)
        if buque_seleccionado.lower() != "buque" and buque_seleccionado.lower() not in file_name_parts:
            file_name_parts.append(buque_seleccionado)

        file_name = "_".join(file_name_parts) + ".xlsx"

        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Guardar archivo Excel",
            file_name,
            "Excel Files (*.xlsx);;All Files (*)"
        )

        if file_path:
            wb = Workbook()
            ws = wb.active

            cell = ws.cell(row=1, column=1, value="VESSEL")
            cell.font = Font(bold=True)

            cell = ws.cell(row=1, column=2)
            cell.value = buque_seleccionado if buque_seleccionado.lower() != "buque" else "Todos"
            cell.font = Font(bold=True)

            cell = ws.cell(row=2, column=1)
            cell.value = "HOTEL"
            cell.font = Font(bold=True)

            cell = ws.cell(row=2, column=2)
            cell.value = hotel_seleccionado if hotel_seleccionado.lower() != "hotel" else "Todos"
            cell.font = Font(bold=True)

            if self.check_fecha.isChecked():
                fecha_inicio = self.date_start1.date().toPyDate()
                fecha_fin = datetime.combine(self.date_end1.date().toPyDate(), time.max)
            else:
                fecha_inicio = None
                fecha_fin = None

            cell = ws.cell(row=3, column=1)
            cell.value = "FECHA"
            cell.font = Font(bold=True)

            cell = ws.cell(row=3, column=2)
            if fecha_inicio is not None:
                cell.value = f"{fecha_inicio} - {fecha_fin}"
            else:
                cell.value = ""

            # Escribir el texto antes de la tabla
            ws.cell(row=5, column=1, value="Informe roomlist").font = Font(bold=True)

            header_fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')
            data_fill = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')

            for col_num, col_name in enumerate(df.columns, 1):
                cell = ws.cell(row=6, column=col_num)
                cell.value = col_name
                cell.fill = header_fill

            for row_num, row_data in enumerate(df.values, start=7):
                for col_num, cell_value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.value = cell_value
                    cell.fill = data_fill

                    # Centrar el número correlativo
                    if col_num == 1:  # Columna "Nro"
                        cell.alignment = Alignment(horizontal="center")

            # Ajustar automáticamente el ancho de las columnas
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter  # Get the column name
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = max_length + 1
                ws.column_dimensions[column].width = adjusted_width

            wb.save(file_path)

    def volver_a_opciones_programar(self):
        self.main_window.stacked_widget.setCurrentIndex(self.main_window.opciones_programar_index)

class TransportesScreen(QWidget):
    def __init__(self, main_window):
        super().__init__()
        self.main_window = main_window
        self.setup_ui()

    def setup_ui(self):
        layout = QVBoxLayout(self)

        # Cuadro de selección de ciudad dentro de AsistenciasScreen
        self.combo_ciudades = QComboBox()
        self.combo_ciudades.addItems(["SCL", "PUQ", "WPU"])  # Agrega las ciudades al combo box
        layout.addWidget(self.combo_ciudades)

        self.check_fecha = QCheckBox("Habilitar filtro por fechas")
        self.check_fecha.setChecked(False)  # Inicialmente deshabilitado
        self.check_fecha.stateChanged.connect(self.actualizar_datos)  # Conectar evento de cambio de estado
        layout.addWidget(self.check_fecha)

        # Layout horizontal para las fechas
        layout_filtro = QHBoxLayout()

        # Selector de fecha de inicio
        self.date_start1 = QDateEdit()
        self.date_start1.setCalendarPopup(True)
        self.date_start1.setDate(QDate.currentDate())
        layout_filtro.addWidget(QLabel("Fecha inicio:"))
        layout_filtro.addWidget(self.date_start1)

        # Selector de fecha de fin
        self.date_end1 = QDateEdit()
        self.date_end1.setCalendarPopup(True)
        self.date_end1.setDate(QDate.currentDate())
        layout_filtro.addWidget(QLabel("Fecha fin:"))
        layout_filtro.addWidget(self.date_end1)

        layout.addLayout(layout_filtro)

        # Label para mostrar asistencias
        self.label = QLabel()  # Mover el label aquí para que sea un atributo de la clase
        layout.addWidget(self.label)

        # Tabla para mostrar datos
        self.table_widget = QTableWidget()
        layout.addWidget(self.table_widget)

        # Botón para generar el Excel
        button_generar_excel = QPushButton("Generar excel")
        button_generar_excel.clicked.connect(self.generar_excel_con_ciudad)  # Conectar al método de generación de Excel
        layout.addWidget(button_generar_excel)

        # Botón "Volver" para regresar a la pantalla anterior (Generación de Reportes)
        button_volver = QPushButton("Volver")
        button_volver.setFixedWidth(100)
        button_volver.clicked.connect(self.volver_a_opciones_programar)
        layout.addWidget(button_volver, alignment=Qt.AlignmentFlag.AlignLeft)

        # Conectar el cambio en el QComboBox a un método
        self.combo_ciudades.currentTextChanged.connect(self.actualizar_datos)
        self.date_start1.dateChanged.connect(self.actualizar_datos)
        self.date_end1.dateChanged.connect(self.actualizar_datos)

        # Actualizar datos inicialmente
        self.actualizar_datos()

    def generar_excel_con_ciudad(self):
        ciudad_seleccionada = self.combo_ciudades.currentText()  # Obtener la ciudad seleccionada
        self.generar_excel(ciudad_seleccionada)  # Llamar a generar_excel con la ciudad seleccionada

    def actualizar_datos(self):
        ciudad_seleccionada = self.combo_ciudades.currentText()  # Obtiene la ciudad seleccionada
        self.label.setText(f"Requerimiento transportes en {ciudad_seleccionada}")  # Actualiza el label

        # Cargar datos en la tabla
        self.cargar_datos(ciudad_seleccionada)

    def cargar_datos(self, ciudad_seleccionada):
        session = get_db_session()
        ciudad_seleccionada = CITY_AIRPORT_CODES.get(ciudad_seleccionada)

        fecha_inicio = self.date_start1.date().toPyDate()
        fecha_fin = datetime.combine(self.date_end1.date().toPyDate(), time.max)

        # Construir la consulta de transporte
        transporte_necesario = (
            session.query(
                Tripulante.tripulante_id,
                Tripulante.estado.label("Estado"),
                Tripulante.nombre.label("First_Name"),
                Tripulante.apellido.label("Last_Name"),
                Tripulante.nacionalidad.label("Nacionalidad"),
                Transporte.nombre.label("Ciudad_Transporte"),
                TripulanteTransporte.fecha.label("Fecha")
            )
            .join(TripulanteTransporte, Tripulante.tripulante_id == TripulanteTransporte.tripulante_id)
            .filter(Transporte.transporte_id == TripulanteTransporte.transporte_id)
            .filter(func.lower(TripulanteTransporte.ciudad) == ciudad_seleccionada.lower())
        )

        if self.check_fecha.isChecked():
            transporte_necesario = transporte_necesario.filter(
                TripulanteTransporte.fecha >= fecha_inicio,
                TripulanteTransporte.fecha <= fecha_fin
            )

        transporte_necesario = transporte_necesario.all()

        ciudad_select = CITY_TO_AIRPORT_CODES.get(ciudad_seleccionada)

        hotel_necesario = (
            session.query(
                Tripulante.tripulante_id,
                Tripulante.estado.label("Estado"),
                Tripulante.nombre.label("First_Name"),
                Tripulante.apellido.label("Last_Name"),
                Hotel.ciudad.label("Ciudad_Hotel"),
                Hotel.nombre.label("Nombre_Hotel")
            )
            .join(TripulanteHotel, Tripulante.tripulante_id == TripulanteHotel.tripulante_id)
            .filter(Hotel.hotel_id == TripulanteHotel.hotel_id)
            .filter(or_(
                func.lower(Hotel.ciudad) == f"hotel {ciudad_select.lower()}",
                func.lower(Hotel.ciudad) == f"day use {ciudad_select.lower()}"
            ))
            .all()
        )

        vuelo_necesario = (
            session.query(
                Tripulante.tripulante_id,
                Tripulante.estado.label("Estado"),
                Tripulante.nombre.label("First_Name"),
                Tripulante.apellido.label("Last_Name"),
                Vuelo.codigo.label("Codigo"),
                Vuelo.fecha.label("Fecha"),
                Vuelo.hora_salida.label("Hora_Salida"),
                Vuelo.hora_llegada.label("Hora_Llegada"),
                Vuelo.aeropuerto_salida.label("Aeropuerto_Salida"),
                Vuelo.aeropuerto_llegada.label("Aeropuerto_Llegada")
            )
            .join(TripulanteVuelo, Tripulante.tripulante_id == TripulanteVuelo.tripulante_id)
            .filter(Vuelo.vuelo_id == TripulanteVuelo.vuelo_id)
            .filter(or_(
                func.lower(Vuelo.aeropuerto_salida) == ciudad_seleccionada.lower(),
                func.lower(Vuelo.aeropuerto_llegada) == ciudad_seleccionada.lower()
            ))
            .all()
        )

        buque_necesario = (
            session.query(
                Tripulante.tripulante_id,
                Buque.nombre.label("Nombre_Buque"),
                EtaCiudad.eta.label("Eta")
            )
            .join(EtaCiudad, Tripulante.tripulante_id == EtaCiudad.tripulante_id)
            .join(Buque, EtaCiudad.buque_id == Buque.buque_id)
            .filter(Buque.buque_id == EtaCiudad.buque_id)
            .all()
        )

        #print(buque_necesario)

        # Organizar vuelos y transportes por tripulante_id
        transporte_dict = defaultdict(list)
        for transporte in transporte_necesario:
            transporte_dict[transporte.tripulante_id].append(transporte)

        vuelo_dict = defaultdict(list)
        for vuelo in vuelo_necesario:
            vuelo_dict[vuelo.tripulante_id].append(vuelo)

        hotel_dict = {hotel.tripulante_id: hotel for hotel in hotel_necesario}
        buque_dict = {buque.tripulante_id: (buque.Nombre_Buque, buque.Eta) for buque in buque_necesario}

        headers = ["Estado", "Transporte", "Hotel Ciudad", "Nombre Hotel", "Código Vuelo Llegada", "Date Llegada", "Hora Llegada", "Hora Pick Up", "Lugar Pick Up", "Código Vuelo Salida", "Date Salida", "Fecha Salida", "Nave", "ETA", "First Name", "Last Name", "Nacionalidad"]
        self.table_widget.setColumnCount(len(headers))
        self.table_widget.setHorizontalHeaderLabels(headers)
        self.table_widget.setRowCount(0)

        # Construir filas para cada tripulante con vuelos y transportes
        for tripulante_id, transportes in transporte_dict.items():
            vuelos = vuelo_dict.get(tripulante_id, [])
            hotel = hotel_dict.get(tripulante_id)
            buque, eta = buque_dict.get(tripulante_id, ("", ""))

            # Filtrar y asociar transporte al vuelo correcto
            for transporte in transportes:
                if 'Ato-Hotel' in transporte.Ciudad_Transporte:  # Transporte hacia el hotel
                    vuelos_llegada = [v for v in vuelos if v.Aeropuerto_Llegada.lower() == ciudad_seleccionada.lower()]
                    for vuelo in vuelos_llegada:
                        row = self.table_widget.rowCount()
                        self.table_widget.insertRow(row)

                        a1 = CITY_TO_AIRPORT_CODES.get(str(vuelo.Aeropuerto_Salida))
                        a2 = CITY_TO_AIRPORT_CODES.get(str(vuelo.Aeropuerto_Llegada))

                        codigo = f"{str(vuelo.Codigo)} {a1}-{a2}"
                        lugar_pick_up = "Ato"

                        self.table_widget.setItem(row, 0, QTableWidgetItem(transporte.Estado))
                        self.table_widget.setItem(row, 1, QTableWidgetItem(transporte.Ciudad_Transporte))
                        #self.table_widget.setItem(row, 2, QTableWidgetItem(str(transporte.Fecha)))
                        self.table_widget.setItem(row, 2, QTableWidgetItem(hotel.Ciudad_Hotel))
                        self.table_widget.setItem(row, 3, QTableWidgetItem(hotel.Nombre_Hotel))
                        self.table_widget.setItem(row, 4, QTableWidgetItem(codigo))
                        self.table_widget.setItem(row, 5, QTableWidgetItem(str(vuelo.Fecha.date())))
                        self.table_widget.setItem(row, 6, QTableWidgetItem(str(vuelo.Hora_Llegada.time())))
                        self.table_widget.setItem(row, 7, QTableWidgetItem(str(vuelo.Hora_Llegada.time())))
                        self.table_widget.setItem(row, 8, QTableWidgetItem(lugar_pick_up))
                        #self.table_widget.setItem(row, 7, QTableWidgetItem(aeropuerto_llegada))
                        self.table_widget.setItem(row, 12, QTableWidgetItem(buque))
                        self.table_widget.setItem(row, 13, QTableWidgetItem(str(eta)))
                        self.table_widget.setItem(row, 14, QTableWidgetItem(transporte.First_Name))
                        self.table_widget.setItem(row, 15, QTableWidgetItem(transporte.Last_Name))
                        self.table_widget.setItem(row, 16, QTableWidgetItem(transporte.Nacionalidad))

                elif 'Hotel-Ato' in transporte.Ciudad_Transporte:  # Transporte desde el hotel
                    vuelos_salida = [v for v in vuelos if v.Aeropuerto_Salida.lower() == ciudad_seleccionada.lower()]
                    for vuelo in vuelos_salida:
                        row = self.table_widget.rowCount()
                        self.table_widget.insertRow(row)

                        a1 = CITY_TO_AIRPORT_CODES.get(str(vuelo.Aeropuerto_Salida))
                        a2 = CITY_TO_AIRPORT_CODES.get(str(vuelo.Aeropuerto_Llegada))

                        if a1 == 'SCL' and a2 == 'PUQ':
                            a1 = f"{a1} Nacional"
                        elif a1 == 'SCL' and a2 != 'PUQ':
                            a1 = f"{a1} Internacional"

                        codigo = f"{str(vuelo.Codigo)} {a1}-{a2}"
                        lugar_pick_up = "Hotel"

                        if a1 == "PUQ":
                            tiempo_a_restar = timedelta(hours=2, minutes=30)
                        elif a1 == "SCL Nacional":
                            tiempo_a_restar = timedelta(hours=2, minutes=30)
                        elif a1 == "SCL Internacional":
                            tiempo_a_restar = timedelta(hours=3, minutes=30)
                        elif a1 == "WPU":
                            tiempo_a_restar = timedelta(hours=1, minutes=30)
                        elif a1 == "KGI":
                            tiempo_a_restar = timedelta(hours=3, minutes=30)
                        else:
                            tiempo_a_restar = timedelta()

                        hora_pick_up = (vuelo.Hora_Salida - tiempo_a_restar).time()

                        self.table_widget.setItem(row, 0, QTableWidgetItem(transporte.Estado))
                        self.table_widget.setItem(row, 1, QTableWidgetItem(transporte.Ciudad_Transporte))
                        #self.table_widget.setItem(row, 2, QTableWidgetItem(str(transporte.Fecha)))
                        self.table_widget.setItem(row, 2, QTableWidgetItem(hotel.Ciudad_Hotel))
                        self.table_widget.setItem(row, 3, QTableWidgetItem(hotel.Nombre_Hotel))
                        self.table_widget.setItem(row, 9, QTableWidgetItem(codigo))
                        self.table_widget.setItem(row, 10, QTableWidgetItem(str(vuelo.Fecha.date())))
                        self.table_widget.setItem(row, 11, QTableWidgetItem(str(vuelo.Hora_Salida.time())))
                        self.table_widget.setItem(row, 7, QTableWidgetItem(str(hora_pick_up)))
                        self.table_widget.setItem(row, 8, QTableWidgetItem(lugar_pick_up))
                        #self.table_widget.setItem(row, 7, QTableWidgetItem(aeropuerto_llegada))
                        self.table_widget.setItem(row, 12, QTableWidgetItem(buque))
                        self.table_widget.setItem(row, 13, QTableWidgetItem(str(eta)))
                        #self.table_widget.setItem(row, 13, QTableWidgetItem(buque.ETA))
                        self.table_widget.setItem(row, 14, QTableWidgetItem(transporte.First_Name))
                        self.table_widget.setItem(row, 15, QTableWidgetItem(transporte.Last_Name))
                        self.table_widget.setItem(row, 16, QTableWidgetItem(transporte.Nacionalidad))

    def generar_excel(self, ciudad_seleccionada):
    # Crear un DataFrame con los datos de la tabla
        data = []
        for row in range(self.table_widget.rowCount()):
            row_data = []
            for column in range(self.table_widget.columnCount()):
                item = self.table_widget.item(row, column)
                row_data.append(item.text() if item else "")
            data.append(row_data)

        # Definir los nombres de las columnas
        column_names = ["Estado", "Transporte", "Hotel Ciudad", "Nombre Hotel", "Código Vuelo Llegada", "Date Llegada", "Hora Llegada", "Hora Pick Up", "Lugar Pick Up", "Código Vuelo Salida", "Date Salida", "Fecha Salida", "Nave", "ETA", "First Name", "Last Name", "Nacionalidad"]

        df = pd.DataFrame(data, columns=column_names)

        # Convertir "Hora Pick Up" a datetime para una ordenación correcta
        df['Hora Pick Up'] = pd.to_datetime(df['Hora Pick Up'], format='%H:%M:%S', errors='coerce')

        # Ordenar el DataFrame por "Hora Pick Up", "Lugar Pick Up" y "Nombre Hotel"
        df = df.sort_values(by=["Hora Pick Up", "Lugar Pick Up", "Nombre Hotel"])

        # Convertir "Hora Pick Up" de nuevo a solo hora para el Excel
        df['Hora Pick Up'] = df['Hora Pick Up'].dt.strftime('%H:%M:%S')

        # Construir el nombre del archivo Excel
        file_name = f'requerimiento_transportes_{ciudad_seleccionada}.xlsx'
        
        # Guardar en un archivo Excel
        file_path, _ = QFileDialog.getSaveFileName(
            self, 
            "Guardar archivo Excel", 
            file_name,  # Usar el nombre del archivo predefinido
            "Excel Files (*.xlsx);;All Files (*)"
        )
        
        if file_path:
            # Crear un nuevo libro de trabajo de Excel
            wb = Workbook()
            ws = wb.active

            # Definir el color de relleno para los encabezados (celeste claro)
            header_fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')
            # Definir el color de relleno para los datos (amarillo claro)
            data_fill = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')

            cell = ws.cell(row=1, column=1)
            cell.value = "Date Pick Up"

            if self.check_fecha.isChecked():
                fecha_inicio = self.date_start1.date().toPyDate()
                fecha_fin = datetime.combine(self.date_end1.date().toPyDate(), time.max)
            else:
                fecha_inicio = None
                fecha_fin = None

            cell = ws.cell(row=1, column=2)
            if fecha_inicio != None:
                cell.value = f"({fecha_inicio}) - ({fecha_fin.date()})"
            else:
                cell.value = ""

            cell = ws.cell(row=2, column=1)
            cell.value = "Ciudad"
            cell = ws.cell(row=2, column=2)
            cell.value = CITY_AIRPORT_CODES.get(ciudad_seleccionada)

            # Escribir los encabezados del DataFrame manualmente
            for col_num, col_name in enumerate(column_names, 1):
                cell = ws.cell(row=5, column=col_num)
                cell.value = col_name
                cell.fill = header_fill  # Aplicar color a los encabezados

            # Escribir los datos del DataFrame y aplicar color a las celdas
            for row_num, row_data in enumerate(df.values, start=6):
                for col_num, cell_value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.value = cell_value
                    cell.fill = data_fill  # Aplicar color a los datos

            # Guardar el archivo Excel con colores aplicados
            wb.save(file_path)

    def volver_a_opciones_programar(self):
        self.main_window.stacked_widget.setCurrentIndex(self.main_window.opciones_programar_index)

class AsistenciasScreen(QWidget):
    def __init__(self, main_window):
        super().__init__()
        self.main_window = main_window
        self.setup_ui()

    def setup_ui(self):
        layout = QVBoxLayout(self)

        # Cuadro de selección de ciudad dentro de AsistenciasScreen
        self.combo_ciudades = QComboBox()
        self.combo_ciudades.addItems(["SCL", "PUQ", "WPU"])  # Agrega las ciudades al combo box
        layout.addWidget(self.combo_ciudades)

        # Label para mostrar asistencias
        self.label = QLabel()  # Mover el label aquí para que sea un atributo de la clase
        layout.addWidget(self.label)

        # Tabla para mostrar datos
        self.table_widget = QTableWidget()
        layout.addWidget(self.table_widget)

        # Botón para generar el Excel
        button_generar_excel = QPushButton("Generar excel")
        button_generar_excel.clicked.connect(self.generar_excel_con_ciudad)  # Conectar al método de generación de Excel
        layout.addWidget(button_generar_excel)

        # Botón "Volver" para regresar a la pantalla anterior (Generación de Reportes)
        button_volver = QPushButton("Volver")
        button_volver.setFixedWidth(100)
        button_volver.clicked.connect(self.volver_a_opciones_programar)
        layout.addWidget(button_volver, alignment=Qt.AlignmentFlag.AlignLeft)

        # Conectar el cambio en el QComboBox a un método
        self.combo_ciudades.currentTextChanged.connect(self.actualizar_datos)

        # Actualizar datos inicialmente
        self.actualizar_datos()

    def generar_excel_con_ciudad(self):
        ciudad_seleccionada = self.combo_ciudades.currentText()  # Obtener la ciudad seleccionada
        self.generar_excel(ciudad_seleccionada)  # Llamar a generar_excel con la ciudad seleccionada

    def actualizar_datos(self):
        ciudad_seleccionada = self.combo_ciudades.currentText()  # Obtiene la ciudad seleccionada
        self.label.setText(f"Asistencias en {ciudad_seleccionada}")  # Actualiza el label

        # Cargar datos en la tabla
        self.cargar_datos(ciudad_seleccionada)

    def cargar_datos(self, ciudad_seleccionada):
        session = get_db_session()  # Obtener la sesión de la base de datos

        # Obtener el nombre de la ciudad a partir del código
        codigo_ciudad = ciudad_seleccionada
        ciudad_seleccionada = CITY_AIRPORT_CODES.get(ciudad_seleccionada).lower()  # Convertir a minúsculas

        # Obtener datos de vuelos de arribo
        arribo_vuelos = (
            session.query(
                Buque.nombre.label("Vessel"),
                Vuelo.aeropuerto_llegada.label("Aeropuerto_Llegada"),
                EtaCiudad.eta.label("ETA"),
                Vuelo.hora_llegada.label("Hora_Arribo"),
                Tripulante.nombre.label("First_Name"),
                Tripulante.apellido.label("Last_Name"),
                Tripulante.condicion.label("Condition"),
                Tripulante.estado.label("Type"),
                Vuelo.codigo.label("Nro_Vuelo_Arribo"),
                Tripulante.tripulante_id,
                Vuelo.fecha.label("Fecha_Vuelo_Arribo")
            )
            .outerjoin(TripulanteVuelo, TripulanteVuelo.tripulante_id == Tripulante.tripulante_id)
            .outerjoin(Vuelo, TripulanteVuelo.vuelo_id == Vuelo.vuelo_id)
            .outerjoin(Buque, Tripulante.buque_id == Buque.buque_id)
            .outerjoin(EtaCiudad, Buque.buque_id == EtaCiudad.buque_id)
            .filter(func.lower(Vuelo.aeropuerto_llegada) == ciudad_seleccionada)
            .all()
        )
        
        # Obtener datos de vuelos de salida
        salida_vuelos = (
            session.query(
                Vuelo.aeropuerto_salida.label("Aeropuerto_Salida"),
                Vuelo.hora_salida.label("Hora_Salida"),
                Vuelo.codigo.label("Nro_Vuelo_Salida"),
                Vuelo.fecha.label("Fecha_Vuelo_Salida"),
                Tripulante.tripulante_id,
                Tripulante.estado.label("Estado")
            )
            .outerjoin(TripulanteVuelo, TripulanteVuelo.tripulante_id == Tripulante.tripulante_id)
            .outerjoin(Vuelo, TripulanteVuelo.vuelo_id == Vuelo.vuelo_id)
            .filter(func.lower(Vuelo.aeropuerto_salida) == ciudad_seleccionada)
            .all()
        )

        # Construir la consulta de transporte
        transporte_necesario = (
            session.query(
                Tripulante.tripulante_id,
                Tripulante.estado.label("Estado"),
                Transporte.nombre.label("Ciudad_Transporte"),
                TripulanteTransporte.fecha.label("Fecha")
            )
            .join(TripulanteTransporte, Tripulante.tripulante_id == TripulanteTransporte.tripulante_id)
            .filter(Transporte.transporte_id == TripulanteTransporte.transporte_id)
            .filter(func.lower(TripulanteTransporte.ciudad) == codigo_ciudad.lower())
            .all()
        )
        tripulantes_con_transporte = {transporte.tripulante_id for transporte in transporte_necesario}

        # Obtener la información de asistencia y proveedores según la ciudad seleccionada
        if codigo_ciudad == "SCL":
            asistencia_data = (
                session.query(
                    TripulanteAsistencia.tripulante_id,
                    TripulanteAsistencia.necesita_asistencia_scl.label("Necesita_Asistencia"),
                    TripulanteAsistencia.proveedor_scl.label("Proveedor")
                )
                .all()
            )
        elif codigo_ciudad == "PUQ":
            asistencia_data = (
                session.query(
                    TripulanteAsistencia.tripulante_id,
                    TripulanteAsistencia.necesita_asistencia_puq.label("Necesita_Asistencia"),
                    TripulanteAsistencia.proveedor_puq.label("Proveedor")
                )
                .all()
            )
        elif codigo_ciudad == "WPU":
            asistencia_data = (
                session.query(
                    TripulanteAsistencia.tripulante_id,
                    TripulanteAsistencia.necesita_asistencia_wpu.label("Necesita_Asistencia"),
                    TripulanteAsistencia.proveedor_wpu.label("Proveedor")
                )
                .all()
            )

        """
        # Imprimir datos de transporte necesario
        print("Datos de Transporte Necesario:")
        for tripulante in session.query(Tripulante).all():  # Obtener todos los tripulantes
            requiere_transporte = "Sí" if tripulante.tripulante_id in tripulantes_con_transporte else "No"
            print(f"ID Tripulante: {tripulante.tripulante_id}, Estado: {tripulante.estado}, Requiere Transporte: {requiere_transporte}")

        
        # Limpiar la tabla
        self.table_widget.setRowCount(0)
        self.table_widget.setColumnCount(18)  # Número correcto de columnas
        self.table_widget.setHorizontalHeaderLabels(
            ["Vessel", "ETA", "First Name", "Last Name", "Type", "Asistencia", "COLACION", 
            "COMIDAS", "Nro Vuelo Arribo", "Fecha Vuelo Arribo", 
            "Hora Arribo", "Hotel", "Habitación", "Date Pick Up", 
            "Hora Pick Up", "Nro Vuelo Salida", "Fecha Vuelo Salida", 
            "Hora Vuelo Salida"]
        )

        # Combinar y llenar la tabla
        for arribo in arribo_vuelos:
            # Buscar un vuelo de salida correspondiente usando el tripulante_id
            salida = next((s for s in salida_vuelos if s.tripulante_id == arribo.tripulante_id), None)
            
            row = self.table_widget.rowCount()
            self.table_widget.insertRow(row)
            # Datos de arribo
            self.table_widget.setItem(row, 0, QTableWidgetItem(str(arribo.Vessel)))  # Vessel
            self.table_widget.setItem(row, 1, QTableWidgetItem(str(arribo.ETA)))  # ETA
            self.table_widget.setItem(row, 2, QTableWidgetItem(str(arribo.First_Name)))  # First Name
            self.table_widget.setItem(row, 3, QTableWidgetItem(str(arribo.Last_Name)))  # Last Name
            self.table_widget.setItem(row, 4, QTableWidgetItem(str(arribo.Type)))  # Type
            self.table_widget.setItem(row, 8, QTableWidgetItem(str(arribo.Nro_Vuelo_Arribo)))  # Nro Vuelo Arribo
            self.table_widget.setItem(row, 9, QTableWidgetItem(str(arribo.Fecha_Vuelo_Arribo)))  # Fecha Vuelo Arribo
            self.table_widget.setItem(row, 10, QTableWidgetItem(str(arribo.Hora_Arribo)))  # Hora Arribo
            
            # Datos de asistencia
            asistencia = arribo.necesita_asistencia_puq if ciudad_seleccionada == 'puq' else \
                        arribo.necesita_asistencia_scl if ciudad_seleccionada == 'scl' else \
                        arribo.necesita_asistencia_wpu if ciudad_seleccionada == 'wpu' else False
            self.table_widget.setItem(row, 5, QTableWidgetItem("Sí" if asistencia else "No"))  # Asistencia

            # Datos de "Comida" y "Hotel"
            reserva_restaurante = (
                session.query(TripulanteRestaurante)
                .join(Restaurante, TripulanteRestaurante.restaurante_id == Restaurante.restaurante_id)
                .filter(TripulanteRestaurante.tripulante_id == arribo.tripulante_id)
                .filter(func.lower(Restaurante.ciudad) == ciudad_seleccionada)
                .first()
            )
            self.table_widget.setItem(row, 7, QTableWidgetItem("Sí" if reserva_restaurante else "No"))  # Comidas

            # Si hay vuelo de salida, se añaden los detalles
            if salida:
                #print(f"Vuelo de salida encontrado para {arribo.First_Name} {arribo.Last_Name}: {salida.Nro_Vuelo_Salida}, {salida.Fecha_Vuelo_Salida}, {salida.Hora_Salida}")
                self.table_widget.setItem(row, 15, QTableWidgetItem(str(salida.Nro_Vuelo_Salida)))  # Nro Vuelo Salida
                self.table_widget.setItem(row, 16, QTableWidgetItem(str(salida.Fecha_Vuelo_Salida)))  # Fecha Vuelo Salida
                self.table_widget.setItem(row, 17, QTableWidgetItem(str(salida.Hora_Salida)))  # Hora Salida
            #else:
                #print(f"No se encontró vuelo de salida para {arribo.First_Name} {arribo.Last_Name}")
        """
    def generar_excel(self, ciudad_seleccionada):
    # Crear un DataFrame con los datos de la tabla
        data = []
        for row in range(self.table_widget.rowCount()):
            row_data = []
            for column in range(self.table_widget.columnCount()):
                item = self.table_widget.item(row, column)
                row_data.append(item.text() if item else "")
            data.append(row_data)

        # Definir los nombres de las columnas
        column_names = [
            "Vessel", "ETA", "First Name", "Last Name", "Type",
            "Nro Vuelo Arribo", "Fecha Vuelo Arribo", "Hora Arribo",
            "Nro Vuelo Salida", "Fecha Vuelo Salida", "Hora Vuelo Salida"
        ]

        # Convertir a DataFrame
        df = pd.DataFrame(data, columns=column_names)

        # Construir el nombre del archivo Excel
        file_name = f'asistencia_{ciudad_seleccionada}.xlsx'
        
        # Guardar en un archivo Excel
        file_path, _ = QFileDialog.getSaveFileName(
            self, 
            "Guardar archivo Excel", 
            file_name,  # Usar el nombre del archivo predefinido
            "Excel Files (*.xlsx);;All Files (*)"
        )
        
        if file_path:
            # Crear un nuevo libro de trabajo de Excel
            wb = Workbook()
            ws = wb.active

            # Definir el color de relleno para los encabezados (celeste claro)
            header_fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')
            # Definir el color de relleno para los datos (amarillo claro)
            data_fill = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')

            # Escribir los encabezados del DataFrame manualmente
            for col_num, col_name in enumerate(column_names, 1):
                cell = ws.cell(row=5, column=col_num)
                cell.value = col_name
                cell.fill = header_fill  # Aplicar color a los encabezados

            # Escribir los datos del DataFrame y aplicar color a las celdas
            for row_num, row_data in enumerate(df.values, start=6):
                for col_num, cell_value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.value = cell_value
                    cell.fill = data_fill  # Aplicar color a los datos

            # Guardar el archivo Excel con colores aplicados
            wb.save(file_path)

    def volver_a_opciones_programar(self):
        self.main_window.stacked_widget.setCurrentIndex(self.main_window.opciones_programar_index)